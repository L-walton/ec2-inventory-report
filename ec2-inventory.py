#!/usr/bin/env python
"""
#!/usr/bin/env python
   Prerequisites:
       pip install openpyxl boto3

   Usage : python inventory.py AWS-profile-name

   This script create a XLSX report with EC2 instances details report on owned by AWS account.
"""
import awspricing
import boto3
import os
import re
import time
from datetime import datetime, timedelta, timezone
from dateutil.parser import parse
import logging
from botocore.exceptions import ClientError
import pprint
import tempfile
import yaml
import json
import csv
import sys
import string
from boto3.session import Session
from operator import itemgetter
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color

# static globals

os.environ['AWS_PROFILE'] = sys.argv[1]
os.environ['AWS_DEFAULT_REGION'] = "us-west-2"
timestr = time.strftime("%Y%m%d-%H%M%S")
logger = logging.getLogger()
logger.setLevel(logging.INFO)


def monitor_cw(instance_id, region):
    # returns cpu utilization from cloudwatch
    now = datetime.utcnow()
    past = now - timedelta(minutes=10090)
    future = now + timedelta(minutes=10)
    cwclient = boto3.client('cloudwatch', region_name=region)
    results = cwclient.get_metric_statistics(Namespace='AWS/EC2',
                                             MetricName='CPUUtilization',
                                             Dimensions=[{
                                                 'Name': 'InstanceId',
                                                 'Value': instance_id
                                             }],
                                             StartTime=past,
                                             EndTime=future,
                                             Period=86400,
                                             Statistics=['Average'])
    datapoints = results['Datapoints']
    load = ''
    if datapoints:
        last_datapoint = sorted(datapoints, key=itemgetter('Timestamp'))[-1]
        utilization = last_datapoint['Average']
        load = round((utilization), 2)
    return load


def monitor_ec2(region):
    # returns row with ec2 details.
    Now = datetime.now(timezone.utc)
    client = boto3.client('ec2', region_name=region)
    paginator = client.get_paginator('describe_instances')
    response_iterator = paginator.paginate()
    for page in response_iterator:
        for obj in page['Reservations']:
            for instance in obj['Instances']:
                InstanceName = None
                InstanceStackName = None
                Platform = "linux"
                ID = instance['InstanceId']
                State = instance['State']['Name']
                LaunchTime = str(instance['LaunchTime'])
                launch_time = parse(LaunchTime)
                td = Now - launch_time
                lt_in_hours = td.total_seconds() // 3600
                PrivateIP = None
                PublicIPADDDR = None
                if State != 'terminated':
                    for tag in instance["Tags"]:
                        try:
                            if tag["Key"] == 'Name':
                                InstanceName = tag["Value"]

                            if tag["Key"] == 'aws:cloudformation:stack-name':
                                InstanceStackName = tag["Value"]

                        except:
                            print("Tag Error", instance, tag)
                    if InstanceName == None:
                        InstanceName = instance['PublicDnsName']

                    root_device_type = instance['RootDeviceType']
                    image_id = instance['ImageId']

                try:
                    ec2 = boto3.resource('ec2', region_name=region)
                except ClientError as ex:
                    print('ec2')
                    if ex.response['Error']['Message']:
                        error_message = ex.response['Error']['Message']
                        print('setup_resource', error_message)
                try:
                    InstanceDetails = ec2.Instance(instance['InstanceId'])
                except ClientError as ex:
                    print('InstanceDetails')
                    if ex.response['Error']['Message']:
                        error_message = ex.response['Error']['Message']
                        print('Instance Details', error_message)
                try:
                    ec2vol = list()
                    Volumes = InstanceDetails.volumes.all()
                    volume_ids = [v.id for v in Volumes]
                    if volume_ids:
                        for volume_id in volume_ids:
                            Vol = ec2.Volume(id=volume_id)
                            ec2vol.append(Vol.attachments[0][u'Device'])
                            ec2vol.append(Vol.size)
                    else:
                        isii = root_device_type + ': ' + image_id
                        ec2vol.append(isii)

                except ClientError as ex:
                    print('Volume')
                    if ex.response['Error']['Message']:
                        error_message = ex.response['Error']['Message']
                        print('Instance Volume Details', error_message)
                if State == 'running':
                    pprint.pprint(InstanceName)

                if State == 'running':
                    try:
                        for inet in instance['NetworkInterfaces']:
                            if 'Association' in inet:
                                if inet['Association']['PublicIp']:
                                    PublicIPADDDR = inet['Association']['PublicIp']

                            if 'Platform' in instance:
                                Platform = instance['Platform']

                            if 'PrivateIpAddress' in instance:
                                PrivateIP = instance['PrivateIpAddress']
                    except ClientError as ex:
                        print('NetworkInterfaces')
                        if ex.response['Error']['Message']:
                            error_message = ex.response['Error']['Message']
                            print('Instance NetworkInterface', error_message)

                    if PublicIPADDDR == None:
                        try:
                            if instance.get(u'PublicIpAddress'):
                                PublicIPADDDR = instance.get(u'PublicIpAddress')
                            else:
                                PublicIPADDDR = "None"
                        except ClientError as ex:
                            print("Received error: %s", ex, exc_info=True)
                            if ex.response['Error']['Message']:
                                error_message = ex.response['Error']['Message']
                                print('Instance PublicIpAddress',
                                      error_message)
                            else:
                                print("Unexpected error: %s" % e)
                                raise e
                    if PrivateIP == None:
                        try:
                            PrivateIP = instance['PrivateIpAddress']
                        except ClientError as ex:
                            print('Instance PrivateIpAddress', ID)
                            if ex.response['Error']['Message']:
                                error_message = ex.response['Error']['Message']
                                print('Instance PrivateIP', error_message)

                if State == 'running':
                    try:
                        HW = data['compute']['models'][region][
                            instance["InstanceType"]]

                    except:
                        print("Error in data", "Name:", InstanceName,
                              "Region:", region, "Type:",
                              instance["InstanceType"])

                if State == 'running':
                    ec2_offer = awspricing.offer('AmazonEC2')
                    try:
                        on_demand_price = ec2_offer.ondemand_hourly(
                            instance["InstanceType"],
                            operating_system='Linux',
                            region=region)
                    except:
                        print("on demand price error",
                              instance["InstanceType"], region)
                        on_demand_price = 0.000

                    try:
                        reserved_price = ec2_offer.reserved_hourly(
                            instance["InstanceType"],
                            operating_system='Linux',
                            lease_contract_length='3yr',
                            offering_class='convertible',
                            purchase_option='Partial Upfront',
                            region=region)
                    except:
                        reserved_price = 0.000

                    if reserved_price != 0.00:
                        rp_cost_total = round(lt_in_hours * reserved_price, 2)
                    else:
                        rp_cost_total = "NA"

                    if on_demand_price != 0.00:
                        od_cost_total = round(lt_in_hours * on_demand_price, 2)
                    else:
                        od_cost_total = "NA"

                if State == 'running':
                    row = list()
                    row.append(instance['Placement']['AvailabilityZone'])
                    row.append(InstanceName)
                    row.append(InstanceStackName)
                    row.append(ID)
                    row.append(instance["InstanceType"])
                    row.append(Platform)
                    row.append(PublicIPADDDR)
                    row.append(PrivateIP)
                    row.append(State)
                    row.append(Account)
                    row.append(HW['CPU'])
                    row.append(monitor_cw(instance["InstanceId"], region))
                    row.append(HW['ECU'])
                    row.append(HW['memoryGiB'])
                    row.append(round(on_demand_price, 4))
                    row.append(round(reserved_price, 4))
                    row.append(LaunchTime)
                    row.append(od_cost_total)
                    row.append(rp_cost_total)
                    row.extend(ec2vol)
                    ws.append(row)


def get_regions():
    """
        List all AWS region available
        :return: list of regions
    """
    client = boto3.client('ec2')
    regions = [
        region['RegionName'] for region in client.describe_regions()['Regions']
    ]
    return regions


def format_xlsx(ws):
    #
    #       Formating XLSX
    #   TOP Row
    try:
        for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1):
            for cell in rows:
                cell.fill = PatternFill("solid", fgColor="0066a1")
                cell.font = Font(color="00FFFFFF", bold=True)
#   Format entire table
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
    except:
        print("rows", rows)

    # set wrapText
    try:
        alignment = Alignment(wrap_text=True)
        for col in ws.iter_cols(min_row=1, min_col=1, max_col=30):
            for cell in col:
                cell.alignment = alignment
    except:
        print("col", col)


# set weight
    for col in ws.iter_cols(min_row=1, min_col=1, max_col=30):
        max_length = 0
        c = col[0].column
        column = get_column_letter(c)
        for cell in col[1:]:
            cell.border = thin_border
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
                    if ' ' in str(cell.value):
                        max_length = ((len(cell.value) + 2) * 1.4) / 2
            except:
                pass
        adjusted_width = (max_length + 6)
        ws.column_dimensions[column].width = adjusted_width


def init_moniroting():
    """
        Script initialisation
        :retur: None
    """
    # Setting regions
    global region_list
    region_list = get_regions()


if __name__ == '__main__':
    import sys
    logging.basicConfig(level=logging.WARNING)
    try:
        print('Loading price.json')
        with open('price.json') as json_file:
            data = json.load(json_file)
        wb = Workbook()
        # grab the active worksheet
        ws = wb.active
        header = [
            'Placement', 'Name', 'Instance Stack Name','Instance ID', 'Instance Type', 'Platform',
            'Public IP', 'Private IP', 'Instance State', 'AWS Account', 'CPU',
            'CPU Utilization Avg', 'ECU', 'memory GiB', 'On demand price',
            'Reserved price', 'LaunchTime', 'On demand total',
            'Reserved total', 'Volume', 'Size GiB', 'Volume', 'Size GiB',
            'Volume', 'Size GiB', 'Volume', 'Size GiB', 'Volume', 'Size GiB',
            'Volume', 'Size GiB'
        ]
        ws.append(header)
        global Account
        iam = boto3.client("iam")
        paginator = iam.get_paginator('list_account_aliases')
        for response in paginator.paginate():
            Account = "\n".join(response['AccountAliases'])
            # Calling functions for every region
        init_moniroting()
        for region in region_list:
            monitor_ec2(region)
        ws = format_xlsx(ws)
        wb.save("inventory-" + Account + "-" + timestr + ".xlsx")
    except ClientError as ex:
        if ex.response['Error']['Message']:
            error_message = ex.response['Error']['Message']
            print('Genric Error', error_message)
            logger.error(error_message)

    except Exception as err:
        print('Exception', err)
        logger.error(err)
